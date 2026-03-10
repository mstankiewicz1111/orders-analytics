import csv
import io
from math import ceil
from zoneinfo import ZoneInfo

from fastapi import Depends, FastAPI, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from sqlalchemy.orm import Session
from starlette.middleware.sessions import SessionMiddleware

from .auth import (
    is_session_authenticated,
    login_user,
    logout_user,
    require_api_auth,
    verify_credentials,
)
from .db import get_db
from .repositories import (
    count_table_rows,
    count_aggregated_symbol_rows,
    get_aggregated_symbol_rows,
    get_last_sync_info,
    get_latest_sync_run,
    get_table_rows,
)
from .settings import settings
from .sync_service import sync_all


WARSAW_TZ = ZoneInfo("Europe/Warsaw")


def format_dt_pl(dt):
    if not dt:
        return None
    return dt.astimezone(WARSAW_TZ).strftime("%Y-%m-%d %H:%M:%S")


app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=settings.session_secret or "change-me")
templates = Jinja2Templates(directory="app/templates")


@app.get("/", response_class=HTMLResponse)
def root():
    return RedirectResponse(url="/login", status_code=302)


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if is_session_authenticated(request):
        return RedirectResponse(url="/admin", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login", response_class=HTMLResponse)
def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    if not verify_credentials(username, password):
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Nieprawidłowy login lub hasło."},
            status_code=401,
        )

    login_user(request)
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/logout")
def logout(request: Request):
    logout_user(request)
    return RedirectResponse(url="/login", status_code=302)


@app.get("/admin", response_class=HTMLResponse)
def admin_panel(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)

    q = (request.query_params.get("q") or "").strip()
    sort = (request.query_params.get("sort") or "id_asc").strip()
    page = max(int(request.query_params.get("page", 1)), 1)

    grouped_q = (request.query_params.get("grouped_q") or "").strip()
    grouped_sort = (request.query_params.get("grouped_sort") or "symbol_asc").strip()
    grouped_page = max(int(request.query_params.get("grouped_page", 1)), 1)

    per_page = 50

    total = count_table_rows(db, q=q)
    total_pages = max(ceil(total / per_page), 1)
    if page > total_pages:
        page = total_pages

    rows = get_table_rows(db, q=q, sort=sort, page=page, per_page=per_page)

    grouped_total = count_aggregated_symbol_rows(db, q=grouped_q)
    grouped_total_pages = max(ceil(grouped_total / per_page), 1)
    if grouped_page > grouped_total_pages:
        grouped_page = grouped_total_pages

    grouped_rows = get_aggregated_symbol_rows(
        db,
        q=grouped_q,
        sort=grouped_sort,
        page=grouped_page,
        per_page=per_page,
    )

    sync_info = get_last_sync_info(db)

    formatted_last_data_fetch_at = format_dt_pl(sync_info["last_data_fetch_at"])
    formatted_last_run_started_at = (
        format_dt_pl(sync_info["last_run"]["started_at"])
        if sync_info["last_run"]
        else None
    )
    formatted_last_run_finished_at = (
        format_dt_pl(sync_info["last_run"]["finished_at"])
        if sync_info["last_run"] and sync_info["last_run"]["finished_at"]
        else None
    )

    return templates.TemplateResponse(
        "table.html",
        {
            "request": request,
            "rows": rows,
            "q": q,
            "sort": sort,
            "page": page,
            "total_pages": total_pages,
            "total": total,
            "grouped_rows": grouped_rows,
            "grouped_q": grouped_q,
            "grouped_sort": grouped_sort,
            "grouped_page": grouped_page,
            "grouped_total_pages": grouped_total_pages,
            "grouped_total": grouped_total,
            "last_data_fetch_at": formatted_last_data_fetch_at,
            "last_run": sync_info["last_run"],
            "last_run_started_at": formatted_last_run_started_at,
            "last_run_finished_at": formatted_last_run_finished_at,
        },
    )


@app.post("/admin/sync")
def admin_sync(
    request: Request,
    db: Session = Depends(get_db),
):
    require_api_auth(request)
    sync_all(db)
    return RedirectResponse(url="/admin", status_code=303)


@app.get("/admin/export-csv")
def export_csv(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)

    q = (request.query_params.get("q") or "").strip()
    sort = (request.query_params.get("sort") or "id_asc").strip()

    rows = get_table_rows(db, q=q, sort=sort, page=1, per_page=100000)
    sync_info = get_last_sync_info(db)
    formatted_last_data_fetch_at = format_dt_pl(sync_info["last_data_fetch_at"])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow(["Data ostatniego pobrania danych", formatted_last_data_fetch_at or "brak danych"])
    writer.writerow([])

    writer.writerow([
        "ID",
        "symbol-kolor",
        "rozmiar",
        "stan dyspozycyjny (M1)",
        "rezerwacje",
        "Całkowita liczba sprzedanych",
    ])

    for row in rows:
        writer.writerow([
            row["id"],
            row["symbol_kolor"],
            row["size_id"],
            row["m1_stan_dyspozycyjny"],
            row["rezerwacje"],
            row["calkowita_liczba_sprzedanych"],
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="stany_magazynowe.csv"'
        },
    )


@app.get("/admin/export-xlsx")
def export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)

    q = (request.query_params.get("q") or "").strip()
    sort = (request.query_params.get("sort") or "id_asc").strip()

    rows = get_table_rows(db, q=q, sort=sort, page=1, per_page=100000)
    sync_info = get_last_sync_info(db)
    formatted_last_data_fetch_at = format_dt_pl(sync_info["last_data_fetch_at"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Stany"

    ws["A1"] = "Data ostatniego pobrania danych"
    ws["B1"] = formatted_last_data_fetch_at or "brak danych"

    headers = [
        "ID",
        "symbol-kolor",
        "rozmiar",
        "stan dyspozycyjny (M1)",
        "rezerwacje",
        "Całkowita liczba sprzedanych",
    ]
    ws.append([])
    ws.append(headers)

    for row in rows:
        ws.append([
            row["id"],
            row["symbol_kolor"],
            row["size_id"],
            row["m1_stan_dyspozycyjny"],
            row["rezerwacje"],
            row["calkowita_liczba_sprzedanych"],
        ])

    widths = {
        "A": 12,
        "B": 30,
        "C": 16,
        "D": 24,
        "E": 14,
        "F": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="stany_magazynowe.xlsx"'
        },
    )


@app.get("/admin/sync-status")
def admin_sync_status(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    row = get_latest_sync_run(db)
    if not row:
        return {
            "status": "idle",
            "started_at": None,
            "finished_at": None,
            "products_found": 0,
            "batches_processed": 0,
            "rows_written_current": 0,
            "rows_written_history": 0,
            "error_message": None,
        }

    return {
        "status": row["status"],
        "started_at": format_dt_pl(row["started_at"]),
        "finished_at": format_dt_pl(row["finished_at"]) if row["finished_at"] else None,
        "products_found": row["products_found"],
        "batches_processed": row["batches_processed"],
        "rows_written_current": row["rows_written_current"],
        "rows_written_history": row["rows_written_history"],
        "error_message": row["error_message"],
    }


@app.get("/admin/export-grouped-csv")
def export_grouped_csv(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)

    grouped_q = (request.query_params.get("grouped_q") or "").strip()
    grouped_sort = (request.query_params.get("grouped_sort") or "symbol_asc").strip()

    rows = get_aggregated_symbol_rows(
        db,
        q=grouped_q,
        sort=grouped_sort,
        page=1,
        per_page=100000,
    )

    sync_info = get_last_sync_info(db)
    formatted_last_data_fetch_at = format_dt_pl(sync_info["last_data_fetch_at"])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")

    writer.writerow(["Data ostatniego pobrania danych", formatted_last_data_fetch_at or "brak danych"])
    writer.writerow([])

    writer.writerow([
        "symbol-kolor",
        "rozmiar",
        "Łączna liczba sprzedanych",
    ])

    for row in rows:
        writer.writerow([
            row["symbol_kolor"],
            row["size_id"],
            row["laczna_liczba_sprzedanych"],
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="stany_zagregowane.csv"'
        },
    )


@app.get("/admin/export-grouped-xlsx")
def export_grouped_xlsx(
    request: Request,
    db: Session = Depends(get_db),
):
    if not is_session_authenticated(request):
        return RedirectResponse(url="/login?next=/admin", status_code=303)

    grouped_q = (request.query_params.get("grouped_q") or "").strip()
    grouped_sort = (request.query_params.get("grouped_sort") or "symbol_asc").strip()

    rows = get_aggregated_symbol_rows(
        db,
        q=grouped_q,
        sort=grouped_sort,
        page=1,
        per_page=100000,
    )

    sync_info = get_last_sync_info(db)
    formatted_last_data_fetch_at = format_dt_pl(sync_info["last_data_fetch_at"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Zagregowane"

    ws["A1"] = "Data ostatniego pobrania danych"
    ws["B1"] = formatted_last_data_fetch_at or "brak danych"

    ws.append([])
    ws.append([
        "symbol-kolor",
        "rozmiar",
        "Łączna liczba sprzedanych",
    ])

    for row in rows:
        ws.append([
            row["symbol_kolor"],
            row["size_id"],
            row["laczna_liczba_sprzedanych"],
        ])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 26

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="stany_zagregowane.xlsx"'
        },
    )
